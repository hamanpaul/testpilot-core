#!/usr/bin/env python3
"""One-shot wifi_llapi inventory alignment (2026-04-24).

See docs/superpowers/specs/2026-04-24-wifi-llapi-align-missing-rows-design.md
for the full design and acceptance criteria.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import TypedDict

from openpyxl import load_workbook
from ruamel.yaml import YAML

REPO_ROOT = Path(__file__).resolve().parents[3]
TEMPLATE_XLSX = REPO_ROOT / "plugins" / "wifi_llapi" / "reports" / "templates" / "wifi_llapi_template.xlsx"
CASES_DIR = REPO_ROOT / "plugins" / "wifi_llapi" / "cases"
TEMPLATE_YAML = CASES_DIR / "_template.yaml"
THIS_DIR = Path(__file__).resolve().parent
REPORT_MD = THIS_DIR / "inventory_alignment_20260424.md"
REPORT_JSON = THIS_DIR / "inventory_alignment_20260424.json"
_FN_ROW_RE = re.compile(r"^D(\d{3,4})_")
_ALLOWED_DIRTY_PATHS = {
    str(REPORT_MD.relative_to(REPO_ROOT)),
    str(REPORT_JSON.relative_to(REPO_ROOT)),
}

PLAN_RENAMES: list[tuple[str, int, str, str, int, str]] = [
    ("D068_discoverymethodenabled_accesspoint_fils.yaml", 68, "wifi-llapi-D068-discoverymethodenabled-accesspoint-fils", "D066_discoverymethodenabled_accesspoint_fils.yaml", 66, "wifi-llapi-D066-discoverymethodenabled-accesspoint-fils"),
    ("D068_discoverymethodenabled_accesspoint_upr.yaml", 68, "wifi-llapi-D068-discoverymethodenabled-accesspoint-upr", "D067_discoverymethodenabled_accesspoint_upr.yaml", 67, "wifi-llapi-D067-discoverymethodenabled-accesspoint-upr"),
    ("D115_getstationstats_accesspoint.yaml", 115, "wifi-llapi-D115-getstationstats-accesspoint", "D109_getstationstats.yaml", 109, "wifi-llapi-D109-getstationstats"),
    ("D115_getstationstats_active.yaml", 115, "wifi-llapi-D115-getstationstats-active", "D110_getstationstats_active.yaml", 110, "wifi-llapi-D110-getstationstats-active"),
    ("D115_getstationstats_associationtime.yaml", 115, "wifi-llapi-D115-getstationstats-associationtime", "D111_getstationstats_associationtime.yaml", 111, "wifi-llapi-D111-getstationstats-associationtime"),
    ("D115_getstationstats_authenticationstate.yaml", 115, "wifi-llapi-D115-getstationstats-authenticationstate", "D112_getstationstats_authenticationstate.yaml", 112, "wifi-llapi-D112-getstationstats-authenticationstate"),
    ("D115_getstationstats_avgsignalstrength.yaml", 115, "wifi-llapi-D115-getstationstats-avgsignalstrength", "D113_getstationstats_avgsignalstrength.yaml", 113, "wifi-llapi-D113-getstationstats-avgsignalstrength"),
    ("D115_getstationstats_avgsignalstrengthbychain.yaml", 115, "wifi-llapi-D115-getstationstats-avgsignalstrengthbychain", "D114_getstationstats_avgsignalstrengthbychain.yaml", 114, "wifi-llapi-D114-getstationstats-avgsignalstrengthbychain"),
]

PLAN_MOVE: tuple[str, int, str, str, int, str] = (
    "D495_retrycount_ssid_stats_basic.yaml",
    495,
    "wifi-llapi-d495-retrycount-basic",
    "D407_retrycount_ssid_stats.yaml",
    407,
    "wifi-llapi-D407-retrycount",
)
PLAN_METADATA_ONLY: list[tuple[str, int, str, int]] = [
    ("D495_retrycount_ssid_stats_verified.yaml", 362, "wifi-llapi-d495-retrycount-verified", 495)
]
PLAN_DELETES: list[tuple[str, int]] = [
    ("D096_uapsdenable.yaml", 96),
    ("D097_vendorie.yaml", 97),
    ("D100_wmmenable.yaml", 100),
    ("D102_configmethodssupported.yaml", 102),
    ("D106_relaycredentialsenable.yaml", 106),
    ("D474_channel_radio_37.yaml", 474),
]
PLAN_CREATE: dict[str, object] = {
    "filename": "D428_channel_neighbour.yaml",
    "row": 428,
    "id": "wifi-llapi-D428-channel-neighbour",
    "name": "Channel — WiFi.AccessPoint.{i}.Neighbour.{i}.",
    "object": "WiFi.AccessPoint.{i}.Neighbour.{i}.",
    "api": "Channel",
    "hlapi_command": 'ubus-cli "WiFi.AccessPoint.{i}.Neighbour.{i}.Channel=36"',
    "llapi_support": "Support",
    "step1_command": 'echo "replace with actual test command"',
    "step2_command": 'echo "replace with verification command"',
    "verification_command": [
        'ubus-cli "WiFi.?" | sed -n "1,200p"',
        "wl -i wl0 status 2>/dev/null || true",
        "wl -i wl1 status 2>/dev/null || true",
        "wl -i wl2 status 2>/dev/null || true",
    ],
}


class SupportRow(TypedDict):
    object: str
    type: str
    param: str
    hlapi: str


class CaseInfo(TypedDict):
    source_row: int | None
    id: str | None


class PlanValidationError(RuntimeError):
    pass


class PostStateError(RuntimeError):
    pass


class ApplyActionsError(RuntimeError):
    def __init__(self, message: str, partial_actions: list[dict]):
        super().__init__(message)
        self.partial_actions = list(partial_actions)


def load_support_rows(xlsx: Path = TEMPLATE_XLSX) -> dict[int, SupportRow]:
    wb = load_workbook(xlsx, read_only=True)
    ws = wb["Wifi_LLAPI"]
    out: dict[int, SupportRow] = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        if row[0] is None and row[3] is None:
            continue
        e = (row[4] or "").strip() if row[4] else ""
        if e == "Support":
            out[i] = {
                "object": row[0] or "",
                "type": row[1] or "",
                "param": (row[2] or "").strip() if row[2] else "",
                "hlapi": row[3] or "",
            }
    return out


def scan_cases(cases_dir: Path = CASES_DIR) -> dict[str, CaseInfo]:
    yaml_loader = YAML(typ="safe")
    out: dict[str, CaseInfo] = {}
    for f in sorted(cases_dir.glob("*.yaml")):
        if f.name.startswith("_"):
            continue
        d = yaml_loader.load(f)
        if not isinstance(d, dict):
            continue
        sr = (d.get("source") or {}).get("row")
        out[f.name] = {
            "source_row": int(sr) if sr is not None else None,
            "id": d.get("id"),
        }
    return out


def clone_cases(cases: dict[str, CaseInfo]) -> dict[str, CaseInfo]:
    return {name: dict(info) for name, info in cases.items()}


def filename_row(name: str) -> int | None:
    m = _FN_ROW_RE.match(name)
    return int(m.group(1)) if m else None


def _yaml_rt() -> YAML:
    """Round-trip YAML loader/dumper that preserves comments and long lines."""
    y = YAML()
    y.preserve_quotes = True
    y.width = 4096
    return y


def _yaml_safe() -> YAML:
    return YAML(typ="safe")


def _replace_top_level_scalar(text: str, key: str, value: str) -> tuple[str, bool]:
    pattern = re.compile(rf"(?m)^(?P<prefix>{re.escape(key)}:\s*)(?P<value>[^\n#]*?)(?P<suffix>\s*(?:#.*)?)$")

    def repl(match: re.Match[str]) -> str:
        return f"{match.group('prefix')}{value}{match.group('suffix')}"

    updated, count = pattern.subn(repl, text, count=1)
    return updated, count > 0


def _replace_source_row(text: str, new_row: int) -> tuple[str, bool]:
    source_match = re.search(r"(?m)^source:\n(?P<body>(?:^[ \t]+.*(?:\n|$))*)", text)
    if not source_match:
        return text, False

    body = source_match.group("body")
    row_pattern = re.compile(r"(?m)^(?P<indent>[ \t]+row:\s*)(?P<value>[^\n#]*?)(?P<suffix>\s*(?:#.*)?)$")

    def repl(match: re.Match[str]) -> str:
        return f"{match.group('indent')}{new_row}{match.group('suffix')}"

    new_body, count = row_pattern.subn(repl, body, count=1)
    if count:
        return text[: source_match.start("body")] + new_body + text[source_match.end("body") :], True

    indent_match = re.search(r"(?m)^([ \t]+)\S", body)
    indent = indent_match.group(1) if indent_match else "  "
    inserted = f"{indent}row: {new_row}\n"
    return text[: source_match.start("body")] + inserted + body + text[source_match.end("body") :], True


def _edit_metadata_via_ruamel(path: Path, new_row: int | None, new_id: str | None) -> dict[str, list]:
    y = _yaml_rt()
    with path.open() as fh:
        data = y.load(fh)
    changes: dict[str, list] = {}
    if new_id is not None and data.get("id") != new_id:
        changes["id"] = [data.get("id"), new_id]
        data["id"] = new_id
    if new_row is not None:
        source = data.setdefault("source", {})
        old = source.get("row")
        if old != new_row:
            changes["source.row"] = [old, new_row]
            source["row"] = new_row
    if changes:
        with path.open("w") as fh:
            y.dump(data, fh)
    return changes


def _git(args: list[str]) -> None:
    subprocess.run(["git", *args], cwd=REPO_ROOT, check=True)


def _display_path(path: Path) -> str:
    if not path.is_absolute():
        return str(path)
    try:
        return str(path.relative_to(REPO_ROOT))
    except ValueError:
        return path.name


def _edit_metadata(path: Path, new_row: int | None, new_id: str | None) -> dict[str, list]:
    """Edit `id` and `source.row` in place while preserving unrelated file layout."""
    text = path.read_text()
    data = _yaml_safe().load(text)
    changes: dict[str, list] = {}

    if new_id is not None and data.get("id") != new_id:
        changes["id"] = [data.get("id"), new_id]
    if new_row is not None:
        source = data.setdefault("source", {})
        old = source.get("row")
        if old != new_row:
            changes["source.row"] = [old, new_row]
    if not changes:
        return changes

    updated = text
    if "id" in changes:
        updated, changed = _replace_top_level_scalar(updated, "id", new_id)
        if not changed:
            return _edit_metadata_via_ruamel(path, new_row, new_id)
    if "source.row" in changes:
        updated, changed = _replace_source_row(updated, new_row)
        if not changed:
            return _edit_metadata_via_ruamel(path, new_row, new_id)
    path.write_text(updated)
    return changes


def execute_rename(old: str, new: str, new_row: int, new_id: str) -> dict:
    src = CASES_DIR / old
    dst = CASES_DIR / new
    _git(["mv", str(src.relative_to(REPO_ROOT)), str(dst.relative_to(REPO_ROOT))])
    fields = _edit_metadata(dst, new_row, new_id)
    _git(["add", str(dst.relative_to(REPO_ROOT))])
    return {"kind": "rename", "row": new_row, "from": old, "to": new, "fields_changed": fields}


def execute_move(old: str, new: str, new_row: int, new_id: str) -> dict:
    record = execute_rename(old, new, new_row, new_id)
    record["kind"] = "move"
    return record


def execute_metadata_only(filename: str, new_row: int, new_id: str | None) -> dict:
    path = CASES_DIR / filename
    fields = _edit_metadata(path, new_row, new_id)
    if fields:
        _git(["add", str(path.relative_to(REPO_ROOT))])
    return {"kind": "metadata", "row": new_row, "from": filename, "to": filename, "fields_changed": fields}


def execute_delete(filename: str) -> dict:
    path = CASES_DIR / filename
    _git(["rm", str(path.relative_to(REPO_ROOT))])
    return {"kind": "delete", "row": None, "from": filename, "to": None, "fields_changed": {}}


def execute_create_from_template(spec: dict) -> dict:
    src = TEMPLATE_YAML
    dst = CASES_DIR / spec["filename"]
    shutil.copy2(src, dst)
    y = _yaml_rt()
    with dst.open() as fh:
        data = y.load(fh)
    data["id"] = spec["id"]
    data["name"] = spec["name"]
    if "source" not in data:
        data["source"] = {}
    data["source"]["row"] = spec["row"]
    data["source"]["object"] = spec["object"]
    data["source"]["api"] = spec["api"]
    data["hlapi_command"] = spec["hlapi_command"]
    data["llapi_support"] = spec["llapi_support"]
    data["steps"][0]["command"] = spec["step1_command"]
    data["steps"][1]["command"] = spec["step2_command"]
    data["verification_command"] = list(spec["verification_command"])
    with dst.open("w") as fh:
        y.dump(data, fh)
    _git(["add", str(dst.relative_to(REPO_ROOT))])
    return {
        "kind": "create",
        "row": spec["row"],
        "from": "_template.yaml",
        "to": spec["filename"],
        "fields_changed": {
            "id": [None, spec["id"]],
            "source.row": [None, spec["row"]],
            "source.object": [None, spec["object"]],
            "source.api": [None, spec["api"]],
        },
    }


def validate_plan(
    support_rows: dict[int, SupportRow],
    cases: dict[str, CaseInfo],
) -> list[str]:
    """Return list of validation error strings; empty list = valid."""
    errors: list[str] = []

    for old, old_row, old_id, new, new_row, new_id in PLAN_RENAMES:
        if old not in cases:
            errors.append(f"rename source missing: {old}")
        else:
            case = cases[old]
            if case["source_row"] != old_row:
                errors.append(f"rename source row drift: {old}")
            if case["id"] != old_id:
                errors.append(f"rename source id drift: {old}")
        if new in cases:
            errors.append(f"rename target already exists: {new}")
        if new_row not in support_rows:
            errors.append(f"rename new_row {new_row} not in Support set")
        fr = filename_row(new)
        if fr != new_row:
            errors.append(f"rename target filename row {fr} != new_row {new_row} ({new})")
        if not new_id.startswith(f"wifi-llapi-D{new_row:03d}-"):
            errors.append(f"rename new_id does not encode row {new_row}: {new_id}")

    old, old_row, old_id, new, new_row, new_id = PLAN_MOVE
    if old not in cases:
        errors.append(f"move source missing: {old}")
    else:
        case = cases[old]
        if case["source_row"] != old_row:
            errors.append(f"move source row drift: {old}")
        if case["id"] != old_id:
            errors.append(f"move source id drift: {old}")
    if new in cases:
        errors.append(f"move target already exists: {new}")
    if new_row not in support_rows:
        errors.append(f"move new_row {new_row} not in Support set")
    fr = filename_row(new)
    if fr != new_row:
        errors.append(f"move target filename row {fr} != new_row {new_row}")
    if not new_id.startswith(f"wifi-llapi-D{new_row:03d}-"):
        errors.append(f"move new_id does not encode row {new_row}: {new_id}")

    for fname, old_row, old_id, new_row in PLAN_METADATA_ONLY:
        if fname not in cases:
            errors.append(f"metadata-only target missing: {fname}")
        else:
            case = cases[fname]
            if case["source_row"] != old_row:
                errors.append(f"metadata-only source row drift: {fname}")
            if case["id"] != old_id:
                errors.append(f"metadata-only source id drift: {fname}")
        if new_row not in support_rows:
            errors.append(f"metadata-only new_row {new_row} not in Support set")

    for fname, stale_row in PLAN_DELETES:
        if fname not in cases:
            errors.append(f"delete target missing: {fname}")
            continue
        fr = filename_row(fname)
        if fr != stale_row:
            errors.append(f"delete filename row {fr} != stale_row {stale_row} ({fname})")
        if cases[fname]["source_row"] != stale_row:
            errors.append(f"delete source row drift: {fname}")
        if stale_row in support_rows:
            errors.append(f"delete stale row still in Support set: {fname}")

    create = PLAN_CREATE
    if create["filename"] in cases:
        errors.append(f"create target already exists: {create['filename']}")
    if create["row"] not in support_rows:
        errors.append(f"create row {create['row']} not in Support set")
    if not TEMPLATE_YAML.exists():
        errors.append(f"_template.yaml not found at {TEMPLATE_YAML}")
    fr = filename_row(create["filename"])
    if fr != create["row"]:
        errors.append(f"create filename row {fr} != row {create['row']}")

    return errors


def _self_check_plan_counts() -> None:
    assert len(PLAN_RENAMES) == 8, len(PLAN_RENAMES)
    assert len(PLAN_METADATA_ONLY) == 1
    assert len(PLAN_DELETES) == 6
    # net cases delta = -6 (deletes) + 1 (create) = -5; renames/move/metadata are net 0


def _ensure_clean_worktree() -> None:
    proc = subprocess.run(
        ["git", "status", "--porcelain"],
        cwd=REPO_ROOT,
        check=True,
        capture_output=True,
        text=True,
    )
    dirty = []
    for line in proc.stdout.splitlines():
        if not line.strip():
            continue
        path = line[3:]
        if path in _ALLOWED_DIRTY_PATHS:
            continue
        dirty.append(line)
    if dirty:
        raise RuntimeError("worktree is not clean:\n" + "\n".join(dirty))


def _planned_actions() -> list[dict]:
    actions: list[dict] = []
    for old, old_row, old_id, new, new_row, new_id in PLAN_RENAMES:
        actions.append(
            {
                "kind": "rename",
                "row": new_row,
                "from": old,
                "to": new,
                "fields_changed": {
                    "id": [old_id, new_id],
                    "source.row": [old_row, new_row],
                },
            }
        )
    old, old_row, old_id, new, new_row, new_id = PLAN_MOVE
    actions.append(
        {
            "kind": "move",
            "row": new_row,
            "from": old,
            "to": new,
            "fields_changed": {
                "id": [old_id, new_id],
                "source.row": [old_row, new_row],
            },
        }
    )
    for fname, old_row, _old_id, new_row in PLAN_METADATA_ONLY:
        actions.append(
            {
                "kind": "metadata",
                "row": new_row,
                "from": fname,
                "to": fname,
                "fields_changed": {
                    "source.row": [old_row, new_row],
                },
            }
        )
    for fname, _stale_row in PLAN_DELETES:
        actions.append(
            {
                "kind": "delete",
                "row": None,
                "from": fname,
                "to": None,
                "fields_changed": {},
            }
        )
    actions.append(
        {
            "kind": "create",
            "row": PLAN_CREATE["row"],
            "from": "_template.yaml",
            "to": PLAN_CREATE["filename"],
            "fields_changed": {
                "id": [None, PLAN_CREATE["id"]],
                "source.row": [None, PLAN_CREATE["row"]],
                "source.object": [None, PLAN_CREATE["object"]],
                "source.api": [None, PLAN_CREATE["api"]],
            },
        }
    )
    return actions


def project_post_cases(cases: dict[str, CaseInfo]) -> dict[str, CaseInfo]:
    projected = clone_cases(cases)

    for old, _old_row, _old_id, new, new_row, new_id in PLAN_RENAMES:
        projected.pop(old)
        projected[new] = {"source_row": new_row, "id": new_id}

    old, _old_row, _old_id, new, new_row, new_id = PLAN_MOVE
    projected.pop(old)
    projected[new] = {"source_row": new_row, "id": new_id}

    for fname, _old_row, old_id, new_row in PLAN_METADATA_ONLY:
        projected[fname] = {"source_row": new_row, "id": old_id}

    for fname, _stale_row in PLAN_DELETES:
        projected.pop(fname)

    projected[PLAN_CREATE["filename"]] = {
        "source_row": PLAN_CREATE["row"],
        "id": PLAN_CREATE["id"],
    }
    return projected


def summarize_inventory(
    support_rows: dict[int, SupportRow],
    cases: dict[str, CaseInfo],
    *,
    template_exists: bool,
) -> dict:
    total = len(cases)
    incl_template = total + (1 if template_exists else 0)

    liberal_missing_rows: list[int] = []
    coverage: dict[int, str] = {}
    filename_rows = {filename_row(fname) for fname in cases}
    source_rows = {info["source_row"] for info in cases.values()}
    for fname, info in cases.items():
        sr = info["source_row"]
        if sr in support_rows and filename_row(fname) == sr:
            coverage.setdefault(sr, fname)

    for row in support_rows:
        if row not in coverage and row not in filename_rows and row not in source_rows:
            liberal_missing_rows.append(row)

    return {
        "total_cases": total,
        "incl_template": incl_template,
        "support_rows": len(support_rows),
        "canonical_coverage": len(coverage),
        "liberal_missing": len(liberal_missing_rows),
        "liberal_missing_rows": liberal_missing_rows,
    }


def _apply_actions() -> list[dict]:
    actions: list[dict] = []
    try:
        for old, _old_row, _old_id, new, new_row, new_id in PLAN_RENAMES:
            actions.append(execute_rename(old, new, new_row, new_id))
        old, _old_row, _old_id, new, new_row, new_id = PLAN_MOVE
        actions.append(execute_move(old, new, new_row, new_id))
        for fname, _old_row, _old_id, new_row in PLAN_METADATA_ONLY:
            actions.append(execute_metadata_only(fname, new_row, None))
        for fname, _stale_row in PLAN_DELETES:
            actions.append(execute_delete(fname))
        actions.append(execute_create_from_template(PLAN_CREATE))
    except ApplyActionsError:
        raise
    except Exception as exc:
        raise ApplyActionsError(str(exc), actions) from exc
    return actions


def _report_sections(actions: list[dict]) -> list[tuple[str, list[dict]]]:
    return [
        ("Renames (8)", [action for action in actions if action["kind"] == "rename"]),
        ("Move + Metadata Fix (2)", [action for action in actions if action["kind"] in ("move", "metadata")]),
        ("Deletes (6)", [action for action in actions if action["kind"] == "delete"]),
        ("New from _template.yaml (1)", [action for action in actions if action["kind"] == "create"]),
    ]


def _format_field_changes(fields_changed: dict[str, list]) -> str:
    if not fields_changed:
        return "—"
    parts = []
    for key, values in fields_changed.items():
        before, after = values
        parts.append(f"`{key}`: {before!r} → {after!r}")
    return "<br>".join(parts)


def write_markdown_report(mode: str, actions: list[dict], post_state: dict | None) -> Path:
    generated_at = datetime.now(timezone.utc).isoformat()
    lines = [
        "# wifi_llapi inventory alignment report",
        "",
        f"- generated_at: `{generated_at}`",
        f"- mode: `{mode}`",
        f"- actions: `{len(actions)}`",
        "",
    ]
    for title, items in _report_sections(actions):
        lines.extend([f"## {title}", ""])
        if not items:
            lines.extend(["(none)", ""])
            continue
        lines.extend(["| row | from | to | fields_changed |", "| --- | --- | --- | --- |"])
        for action in items:
            lines.append(
                f"| {action['row'] or '—'} | `{action['from']}` | `{action['to'] or '—'}` | {_format_field_changes(action['fields_changed'])} |"
            )
        lines.append("")
    lines.extend(["## Post state", ""])
    if post_state is None:
        lines.append("_not-run_")
    else:
        lines.extend(["```json", json.dumps(post_state, indent=2, ensure_ascii=False, sort_keys=True), "```"])
    REPORT_MD.write_text("\n".join(lines) + "\n")
    return REPORT_MD


def write_json_report(mode: str, actions: list[dict], post_state: dict | None) -> Path:
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "mode": mode,
        "actions": actions,
        "post_state": post_state,
    }
    REPORT_JSON.write_text(json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=True) + "\n")
    return REPORT_JSON


def verify_post_state(expected_state: dict | None = None) -> dict:
    """Re-scan repo and assert acceptance criteria. Raises on failure."""
    rows = load_support_rows()
    cases = scan_cases()
    template_exists = TEMPLATE_YAML.exists()
    support_count = len(rows)
    state = summarize_inventory(rows, cases, template_exists=template_exists)
    expected = expected_state or {
        "total_cases": support_count,
        "incl_template": support_count + (1 if template_exists else 0),
        "support_rows": support_count,
        "canonical_coverage": support_count,
        "liberal_missing": 0,
        "liberal_missing_rows": [],
    }
    errors = []
    if state["total_cases"] != expected["total_cases"]:
        errors.append(f"total cases = {state['total_cases']}, expected {expected['total_cases']}")
    if state["incl_template"] != expected["incl_template"]:
        errors.append(f"total incl _template = {state['incl_template']}, expected {expected['incl_template']}")
    if state["canonical_coverage"] != expected["canonical_coverage"]:
        errors.append(
            f"canonical coverage = {state['canonical_coverage']}/{support_count}, expected {expected['canonical_coverage']}/{support_count}"
        )
    if state["liberal_missing_rows"] != expected["liberal_missing_rows"]:
        errors.append(f"liberal-missing rows: {state['liberal_missing_rows']}, expected {expected['liberal_missing_rows']}")
    if errors:
        raise PostStateError("; ".join(errors) + f" | state={state}")
    return state


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Actually mutate the working tree (default: dry-run).",
    )
    args = parser.parse_args(argv)

    _self_check_plan_counts()
    support_rows = load_support_rows()
    cases = scan_cases()
    errors = validate_plan(support_rows, cases)
    if errors:
        print("plan validation failed:", file=sys.stderr)
        for err in errors:
            print(f"- {err}", file=sys.stderr)
        return 1

    mode = "apply" if args.apply else "dry-run"
    print(f"mode={mode} | support_rows={len(support_rows)} | current_cases={len(cases)}")
    actions: list[dict] = []
    post: dict | None = None
    pending_error: Exception | None = None
    if args.apply:
        expected_post_state = summarize_inventory(
            support_rows,
            project_post_cases(cases),
            template_exists=TEMPLATE_YAML.exists(),
        )
        _ensure_clean_worktree()
        try:
            actions = _apply_actions()
            try:
                post = verify_post_state(expected_post_state)
            except Exception as exc:
                pending_error = exc
        except ApplyActionsError as exc:
            actions = exc.partial_actions
            pending_error = exc
    else:
        actions = _planned_actions()
    report_md = write_markdown_report(mode, actions, post)
    report_json = write_json_report(mode, actions, post)
    print(f"actions: {len(actions)}")
    print(f"reports: {_display_path(report_md)}, {_display_path(report_json)}")
    if post is not None:
        print(f"post_state: {json.dumps(post, ensure_ascii=False, sort_keys=True)}")
    if pending_error is not None:
        raise pending_error
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
